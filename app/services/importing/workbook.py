"""Lectura del XLSX.

El archivo se abre **una sola vez**, al subirlo, y se convierte en filas
tipadas. A partir de ahi el importador trabaja contra PostgreSQL: ningun
request posterior vuelve a abrir el libro ni depende de una ruta local.

Las columnas se localizan por nombre de cabecera, no por posicion, y con
alias: si el archivo renombra una columna el importador lo detecta y lo avisa
en vez de leer la columna equivocada en silencio.
"""

from __future__ import annotations

import io
import zipfile
from dataclasses import dataclass, field
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.core.errors import APIError
from app.core.masters import fold
from app.models.importing import ImportEntity


class WorkbookError(APIError):
    status_code = 422
    code = "IMPORT_FILE_INVALID"
    message = "El archivo no es un Excel legible"


#: Cabeceras que identifican cada hoja y como se llama cada campo dentro. La
#: primera coincidencia gana; los alias cubren tildes, mayusculas y la errata
#: "indentificacion" del maestro original.
FIELD_ALIASES: dict[ImportEntity, dict[str, tuple[str, ...]]] = {
    ImportEntity.PRODUCT_CATEGORY: {
        "name": ("CATEGORIA", "NOMBRE"),
        "parent": ("CATEGORIA PADRE", "PADRE"),
        "display_path": ("NOMBRE A MOSTRAR", "RUTA"),
        "expense_account": ("CUENTA DE GASTO",),
        "income_account": ("CUENTA DE INGRESO",),
        "valuation_account": ("CUENTA DE VALORIZACION",),
        "stock_in_account": ("CUENTA DE ENTRADA DE STOCK",),
        "stock_out_account": ("CUENTA DE SALIDA DE STOCK",),
    },
    ImportEntity.POS_CATEGORY: {
        "name": ("NOMBRE", "CATEGORIA"),
        "parent": ("CATEGORIA PADRE", "PADRE"),
    },
    ImportEntity.PRODUCT: {
        "name": ("NOMBRE",),
        "internal_reference": ("REFERENCIA INTERNA", "REFERENCIA", "CODIGO"),
        "category": ("CATEGORIA DE PRODUCTO", "CATEGORIA"),
        "pos_category": ("CATEGORIA PDV", "CATEGORIA PUNTO DE VENTA"),
        "sale_tax": ("IMPUESTO VENTA", "IMPUESTO DE VENTA"),
        "purchase_tax": ("IMPUESTO DE COMPRA", "IMPUESTO COMPRA"),
        "sellable": ("SE PUEDE VENDER?", "¿SE PUEDE VENDER?", "SE VENDE"),
        "purchasable": ("SE PUEDE COMPRAR?", "¿SE PUEDE COMPRAR?", "SE COMPRA"),
        "available_in_pos": (
            "DISPONIBLE EN PUNTO DE VENTA?",
            "DISPONIBLE EN PUNTO DE VENTA",
            "DISPONIBLE EN POS",
        ),
        "sale_price": ("PRECIO DE VENTA", "PRECIO"),
        "cost": ("COSTO",),
        "uom": ("UNIDAD DE MEDIDA",),
        "purchase_uom": ("UNIDAD DE MEDIDA DE COMPRA", "UNIDAD DE COMPRA"),
    },
    ImportEntity.PARTNER: {
        "name": ("NOMBRE",),
        "document_type": ("TIPO DE IDENTIFICACION", "TIPO DE DOCUMENTO"),
        "document_number": (
            "NUMERO DE INDENTIFICACION",
            "NUMERO DE IDENTIFICACION",
            "NUMERO DE DOCUMENTO",
        ),
        "address": ("CALLE", "DIRECCION"),
        "district": ("DISTRITO",),
        "province": ("PROVINCIA",),
        "department": ("DEPARTAMENTO",),
        "country": ("PAIS",),
        "email": ("CORREO", "EMAIL"),
        "mobile": ("CELULAR",),
        "phone": ("TELEFONO",),
        "account": ("NUMERO DE CUENTA",),
        "bank": ("BANCO",),
    },
    ImportEntity.STOCK: {
        "product": ("PRODUCTO",),
        "uom": ("UNIDAD DE MEDIDA", "UNIDAD"),
        "quantity": ("CANTIDAD",),
        "location": ("UBICACION",),
    },
    ImportEntity.RECIPE: {
        "product": ("NOMBRE DEL PRODUCTO A PREPARAR", "PRODUCTO A PREPARAR"),
        "quantity": ("CANTIDAD",),
        "uom": ("UNIDAD DE MEDIDA DEL PRODUCTO PREPARADO",),
        "component": ("INSUMO",),
        "component_quantity": ("CANTIDAD DEL INSUMO",),
        "component_uom": ("UNIDAD DE MEDIDA DEL INSUMO",),
    },
}

#: Campos que, presentes a la vez, identifican la entidad de una hoja.
ENTITY_SIGNATURES: tuple[tuple[ImportEntity, tuple[str, ...]], ...] = (
    (ImportEntity.RECIPE, ("product", "component", "component_quantity")),
    (ImportEntity.PRODUCT, ("name", "internal_reference")),
    (ImportEntity.PARTNER, ("name", "document_type", "document_number")),
    (ImportEntity.STOCK, ("product", "quantity", "location")),
    (ImportEntity.PRODUCT_CATEGORY, ("name", "parent", "display_path")),
    (ImportEntity.POS_CATEGORY, ("name", "parent")),
)


@dataclass(slots=True)
class WorkbookSheet:
    """Una hoja ya interpretada."""

    name: str
    headers: list[str]
    entity: ImportEntity | None
    #: Indice de columna por campo logico.
    columns: dict[str, int]
    #: ``(numero_de_fila_en_excel, valores)`` solo de las filas con contenido.
    rows: list[tuple[int, list[Any]]]
    warnings: list[str] = field(default_factory=list)

    def value(self, row: list[Any], field_name: str) -> Any:
        index = self.columns.get(field_name)
        if index is None or index >= len(row):
            return None
        value = row[index]
        if isinstance(value, str):
            value = value.strip()
            return value or None
        return value


def _match_columns(headers: list[str], entity: ImportEntity) -> dict[str, int]:
    folded = {fold(header): index for index, header in enumerate(headers) if header}
    columns: dict[str, int] = {}
    for field_name, aliases in FIELD_ALIASES[entity].items():
        for alias in aliases:
            index = folded.get(fold(alias))
            if index is not None:
                columns[field_name] = index
                break
    return columns


def _detect_entity(headers: list[str]) -> tuple[ImportEntity | None, dict[str, int]]:
    for entity, required in ENTITY_SIGNATURES:
        columns = _match_columns(headers, entity)
        if all(name in columns for name in required):
            return entity, columns
    return None, {}


def read_workbook(payload: bytes) -> list[WorkbookSheet]:
    """Convierte el binario en hojas interpretadas.

    Lanza ``WorkbookError`` si el archivo esta corrupto, vacio o no es un XLSX.
    """
    if not payload:
        raise WorkbookError("El archivo esta vacio")
    try:
        workbook = load_workbook(io.BytesIO(payload), data_only=True, read_only=True)
    except (InvalidFileException, zipfile.BadZipFile, KeyError, OSError, ValueError) as exc:
        raise WorkbookError("El archivo no se pudo abrir como Excel") from exc

    sheets: list[WorkbookSheet] = []
    try:
        for name in workbook.sheetnames:
            worksheet = workbook[name]
            iterator = worksheet.iter_rows(values_only=True)
            try:
                header_row = next(iterator)
            except StopIteration:
                sheets.append(
                    WorkbookSheet(
                        name=name,
                        headers=[],
                        entity=None,
                        columns={},
                        rows=[],
                        warnings=["Hoja vacia"],
                    )
                )
                continue

            headers = [str(cell).strip() if cell is not None else "" for cell in header_row]
            while headers and not headers[-1]:
                headers.pop()

            entity, columns = _detect_entity(headers)
            warnings: list[str] = []
            if entity is None and headers:
                warnings.append("No se reconocio ninguna entidad conocida en esta hoja")
            elif entity is not None:
                missing = [name_ for name_ in FIELD_ALIASES[entity] if name_ not in columns]
                if missing:
                    warnings.append("Columnas no encontradas: " + ", ".join(sorted(missing)))

            rows: list[tuple[int, list[Any]]] = []
            width = max(len(headers), 1)
            for number, row in enumerate(iterator, start=2):
                values = list(row[:width])
                if all(
                    value is None or (isinstance(value, str) and not value.strip())
                    for value in values
                ):
                    continue
                rows.append((number, values))

            sheets.append(
                WorkbookSheet(
                    name=name,
                    headers=headers,
                    entity=entity,
                    columns=columns,
                    rows=rows,
                    warnings=warnings,
                )
            )
    finally:
        workbook.close()

    if not sheets:
        raise WorkbookError("El libro no contiene hojas")
    return sheets


def analyze_workbook(sheets: list[WorkbookSheet]) -> list[dict[str, Any]]:
    """Resumen por hoja para el informe previo al mapping."""
    return [
        {
            "name": sheet.name,
            "rows": len(sheet.rows),
            "columns": len(sheet.headers),
            "headers": sheet.headers,
            "entity": sheet.entity.value if sheet.entity else None,
            "detected": len(sheet.rows),
            "warnings": sheet.warnings,
        }
        for sheet in sheets
    ]
